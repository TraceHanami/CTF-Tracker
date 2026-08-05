"""
exporter.py — Generate Excel (.xlsx) and PDF reports for CTF & Hackathon events.

Each event dict is expected to have an optional "status" field:
    "Ongoing"  → user marked it as Continue/Active via the dashboard
    "Upcoming" → normal (default)

Removed events are already filtered out before these functions are called.

generate_excel(events) → str  (absolute filepath)
generate_pdf(events)   → str  (absolute filepath)
"""

from __future__ import annotations
import os
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", "/tmp/data" if os.environ.get("VERCEL") else BASE_DIR / "data"))
try:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    DATA_DIR = Path("/tmp/data")
    DATA_DIR.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _price_str(e: dict) -> str:
    v = e.get("price", 0)
    return "Free" if v == 0 else f"₹{v}"

def _status(e: dict) -> str:
    return e.get("status", "Upcoming")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(events: list[dict], *, teams: list[dict] | None = None) -> str:
    """
    Generate a professionally formatted Excel workbook.

    Sheet 1 — "CTF & Hackathon Events":
      Columns: #, Status, Title, Type, Organizer, Price, Mode, Location, Date, Register
      - Ongoing rows → green left border + light green fill
      - Register column → clickable hyperlink

    Sheet 2 — "Team Rosters" (only when teams data is supplied):
      Columns: #, Team Name, Lead, Members, Member Count, Participating
      - Participating teams → green fill
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    filepath = str((DATA_DIR / "events_export.xlsx").resolve())

    COLUMNS = ["#", "Status", "Title", "Type", "Organizer",
               "Price (INR)", "Mode", "Location", "Date", "Registration Link"]
    NUM_COLS = len(COLUMNS)

    # ── Colours ──────────────────────────────────────────────────────────────
    thin       = Side(border_style="thin",   color="BFBFBF")
    thick_green= Side(border_style="medium", color="16A34A")
    cell_border= Border(left=thin, right=thin, top=thin, bottom=thin)

    FILL_ODD     = PatternFill("solid", fgColor="F0F7FF")
    FILL_EVEN    = PatternFill("solid", fgColor="FFFFFF")
    FILL_ONGOING = PatternFill("solid", fgColor="F0FDF4")   # light green
    HEADER_FILL  = PatternFill("solid", fgColor="2563EB")
    TITLE_FILL   = PatternFill("solid", fgColor="1E3A5F")

    HEADER_FONT  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ONGOING_FONT_STATUS = Font(name="Calibri", size=10, bold=True, color="16A34A")
    UPCOMING_FONT_STATUS= Font(name="Calibri", size=10, color="6B7280")
    LINK_FONT    = Font(name="Calibri", size=10, color="1D4ED8", underline="single")
    DATA_FONT    = Font(name="Calibri", size=10)

    col_last = get_column_letter(NUM_COLS)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "CTF & Hackathon Events"

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_last}1")
    tc = ws["A1"]
    tc.value = (f"CTF & Hackathon Events Report  —  "
                f"Generated {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC")
    tc.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    tc.fill      = TITLE_FILL
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Sub-title ─────────────────────────────────────────────────────────────
    ongoing_count  = sum(1 for e in events if _status(e) == "Ongoing")
    upcoming_count = len(events) - ongoing_count

    ws.merge_cells(f"A2:{col_last}2")
    sc = ws["A2"]
    sc.value = (
        f"Filters: Price ≤ ₹500 or Free  |  Online (Global)  |  Offline: TN, Kerala & Bengaluru  "
        f"|  Total: {len(events)}  |  Ongoing: {ongoing_count}  |  Upcoming: {upcoming_count}"
    )
    sc.font      = Font(name="Calibri", italic=True, size=10, color="4A4A4A")
    sc.fill      = PatternFill("solid", fgColor="D6E4F7")
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Header row (row 3) ────────────────────────────────────────────────────
    for ci, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=ci, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = cell_border
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, e in enumerate(events, start=4):
        status    = _status(e)
        is_ongoing= status == "Ongoing"
        base_fill = FILL_ONGOING if is_ongoing else (FILL_ODD if ri % 2 == 0 else FILL_EVEN)

        link = e.get("registration_link", "")
        row_data = [
            ri - 3,                        # #
            status,                        # Status
            e.get("title", ""),            # Title
            e.get("event_type", ""),       # Type
            e.get("organizer", ""),        # Organizer
            _price_str(e),                 # Price
            e.get("mode", ""),             # Mode
            e.get("location", ""),         # Location
            e.get("date", ""),             # Date
            link,                          # Registration Link
        ]

        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill      = base_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font      = DATA_FONT

            # Left border: thick green for ongoing, thin for others
            if is_ongoing and ci == 1:
                cell.border = Border(left=thick_green, right=thin, top=thin, bottom=thin)
            else:
                cell.border = cell_border

            # Status cell colour
            if ci == 2:
                cell.font = ONGOING_FONT_STATUS if is_ongoing else UPCOMING_FONT_STATUS

            # Hyperlink for Registration Link column
            if ci == NUM_COLS and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font      = LINK_FONT

        ws.row_dimensions[ri].height = 18

    # ── Auto-adjust column widths ──────────────────────────────────────────────
    col_min_max = {
        1: (4,  6),    # #
        2: (10, 14),   # Status
        3: (20, 55),   # Title
        4: (10, 18),   # Type
        5: (16, 44),   # Organizer
        6: (10, 16),   # Price
        7: (8,  14),   # Mode
        8: (16, 46),   # Location
        9: (10, 16),   # Date
        10:(18, 60),   # Registration Link
    }
    for ci, col_name in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(ci)
        max_len    = len(col_name)
        for ri in range(4, 4 + len(events)):
            v = ws.cell(row=ri, column=ci).value or ""
            max_len = max(max_len, len(str(v)))
        mn, mx = col_min_max.get(ci, (10, 50))
        ws.column_dimensions[col_letter].width = max(mn, min(max_len + 3, mx))

    # ── Freeze panes at row 4 ─────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Footer row ────────────────────────────────────────────────────────────
    fr = 4 + len(events) + 1
    ws.merge_cells(f"A{fr}:{col_last}{fr}")
    fc = ws[f"A{fr}"]
    fc.value     = "© CTF & Hackathon Tracker — Automated Report"
    fc.font      = Font(name="Calibri", size=9, italic=True, color="9CA3AF")
    fc.alignment = Alignment(horizontal="right")

    # ── Sheet 2: Team Rosters (if teams data provided) ───────────────────────
    if teams:
        ws2 = wb.create_sheet("Team Rosters")

        T_HEADER_FILL = PatternFill("solid", fgColor="16A34A")   # green-600
        T_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        T_FILL_PART   = PatternFill("solid", fgColor="F0FDF4")   # light green
        T_FILL_ODD    = PatternFill("solid", fgColor="F9FAFB")
        T_FILL_EVEN   = PatternFill("solid", fgColor="FFFFFF")
        T_GREEN_FONT  = Font(name="Calibri", size=10, bold=True, color="15803D")
        T_GREY_FONT   = Font(name="Calibri", size=10, color="6B7280")

        T_COLS = ["#", "Team Name", "Lead", "Members", "Member Count", "Participating"]

        # Title
        ws2.merge_cells("A1:F1")
        t2 = ws2["A1"]
        t2.value = (f"Team Rosters — CTF & Hackathon Event  |  "
                    f"Generated {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC")
        t2.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        t2.fill      = PatternFill("solid", fgColor="14532D")
        t2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 26

        # Summary sub-row
        participating_count = sum(1 for t in teams if t.get("participating"))
        ws2.merge_cells("A2:F2")
        s2 = ws2["A2"]
        s2.value = (f"Total Teams: {len(teams)}  |  "
                    f"Participating: {participating_count}  |  "
                    f"Not Participating: {len(teams) - participating_count}")
        s2.font      = Font(name="Calibri", italic=True, size=10, color="4A4A4A")
        s2.fill      = PatternFill("solid", fgColor="DCFCE7")
        s2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[2].height = 18

        # Header row (row 3)
        for ci, col_name in enumerate(T_COLS, 1):
            c = ws2.cell(row=3, column=ci, value=col_name)
            c.font      = T_HEADER_FONT
            c.fill      = T_HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = cell_border
        ws2.row_dimensions[3].height = 22

        # Data rows
        for ri, t in enumerate(teams, start=4):
            is_part   = bool(t.get("participating"))
            base_fill = T_FILL_PART if is_part else (T_FILL_ODD if ri % 2 == 0 else T_FILL_EVEN)
            members_str = ", ".join(t.get("members") or []) or "—"
            row_data = [
                ri - 3,
                t.get("team", ""),
                t.get("lead", ""),
                members_str,
                len(t.get("members") or []),
                "✅ Yes" if is_part else "—",
            ]
            for ci, value in enumerate(row_data, 1):
                c = ws2.cell(row=ri, column=ci, value=value)
                c.fill      = base_fill
                c.alignment = Alignment(vertical="center", wrap_text=(ci == 4))
                c.border    = cell_border
                if ci == 6:
                    c.font = T_GREEN_FONT if is_part else T_GREY_FONT
                else:
                    c.font = Font(name="Calibri", size=10)
            ws2.row_dimensions[ri].height = 20

        # Footer row
        fr2 = 4 + len(teams) + 1
        ws2.merge_cells(f"A{fr2}:F{fr2}")
        fc2 = ws2[f"A{fr2}"]
        fc2.value     = "© CTF & Hackathon Tracker — Team Rosters Export"
        fc2.font      = Font(name="Calibri", size=9, italic=True, color="9CA3AF")
        fc2.alignment = Alignment(horizontal="right")

        # Column widths
        t_widths = [4, 22, 30, 60, 14, 14]
        for ci, w in enumerate(t_widths, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

        ws2.freeze_panes = "A4"

    wb.save(filepath)
    return filepath


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_pdf(events: list[dict]) -> str:
    """
    Generate a professionally styled landscape-A4 PDF.

    Columns: #, Status, Title, Type, Organizer, Price, Mode, Location, Date, Register
    - Ongoing rows → light-green fill
    - Status column → colour-coded text
    - Registration Link shown as shortened clickable URL
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    filepath = str((DATA_DIR / "events_export.pdf").resolve())

    # ── Colours ───────────────────────────────────────────────────────────────
    BRAND_BLUE   = colors.HexColor("#2563EB")
    BRAND_DARK   = colors.HexColor("#1E3A5F")
    ROW_ODD      = colors.HexColor("#EFF6FF")
    ROW_EVEN     = colors.white
    ROW_ONGOING  = colors.HexColor("#F0FDF4")
    HEADER_BG    = BRAND_BLUE
    HEADER_FG    = colors.white
    GREEN_STATUS = colors.HexColor("#16A34A")
    GREY_STATUS  = colors.HexColor("#6B7280")
    BORDER_COLOR = colors.HexColor("#BFDBFE")
    TEXT_MUTED   = colors.HexColor("#6B7280")
    TEXT_DARK    = colors.HexColor("#111827")
    LINK_COLOR   = colors.HexColor("#1D4ED8")

    # ── Doc setup ─────────────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=18 * mm,  bottomMargin=20 * mm,
        title="CTF & Hackathon Events Report",
        author="CTF Tracker",
    )

    # ── Paragraph styles ──────────────────────────────────────────────────────
    styles   = getSampleStyleSheet()
    s_title  = ParagraphStyle("T", parent=styles["Title"],  fontSize=20,
                               textColor=BRAND_DARK, alignment=TA_CENTER,
                               fontName="Helvetica-Bold", spaceAfter=3)
    s_sub    = ParagraphStyle("S", parent=styles["Normal"], fontSize=8.5,
                               textColor=TEXT_MUTED, alignment=TA_CENTER,
                               fontName="Helvetica-Oblique", spaceAfter=1)
    s_meta   = ParagraphStyle("M", parent=styles["Normal"], fontSize=8,
                               textColor=TEXT_MUTED, alignment=TA_CENTER,
                               fontName="Helvetica")
    s_cell   = ParagraphStyle("C", parent=styles["Normal"], fontSize=7.5,
                               textColor=TEXT_DARK,  fontName="Helvetica", leading=10)
    s_link   = ParagraphStyle("L", parent=s_cell, textColor=LINK_COLOR)
    s_ongoing= ParagraphStyle("ON", parent=s_cell, textColor=GREEN_STATUS,
                               fontName="Helvetica-Bold")
    s_upcoming=ParagraphStyle("UP", parent=s_cell, textColor=GREY_STATUS)
    s_hdr    = ParagraphStyle("H", parent=s_cell,  fontSize=8, textColor=HEADER_FG,
                               fontName="Helvetica-Bold", alignment=TA_CENTER)

    # ── Footer ────────────────────────────────────────────────────────────────
    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(TEXT_MUTED)
        ts = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")
        pw = landscape(A4)[0]
        canvas.drawString(14 * mm, 9 * mm, f"Generated: {ts}")
        canvas.drawCentredString(pw / 2, 9 * mm, "© CTF & Hackathon Tracker — Automated Report")
        canvas.drawRightString(pw - 14 * mm, 9 * mm, f"Page {doc.page}")
        canvas.restoreState()

    # ── Story ─────────────────────────────────────────────────────────────────
    ongoing_count  = sum(1 for e in events if _status(e) == "Ongoing")
    upcoming_count = len(events) - ongoing_count

    story = []
    story.append(Paragraph("🚩 CTF &amp; Hackathon Events Report", s_title))
    story.append(Paragraph(
        "Filters: Price ≤ ₹500 or Free &nbsp;|&nbsp; Online (Global) &nbsp;|&nbsp; "
        "Offline: Tamil Nadu, Kerala &amp; Bengaluru only", s_sub))
    story.append(Paragraph(
        f"Generated {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')} &nbsp;·&nbsp; "
        f"{len(events)} events &nbsp;·&nbsp; "
        f"<font color='#16A34A'><b>{ongoing_count} Ongoing</b></font> &nbsp;·&nbsp; "
        f"{upcoming_count} Upcoming", s_meta))
    story.append(Spacer(1, 3 * mm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=BRAND_BLUE, spaceAfter=3 * mm))

    if not events:
        story.append(Paragraph("No events matched the current filters.", styles["Normal"]))
    else:
        # Column definitions
        COLS   = ["#", "Status", "Title", "Type", "Organizer",
                  "Price", "Mode", "Location", "Date", "Register"]
        WIDTHS = [7*mm, 18*mm, 62*mm, 19*mm, 42*mm,
                  14*mm, 16*mm, 42*mm, 19*mm, 30*mm]

        def _shorten_url(url: str) -> str:
            """Show just the domain + first path segment for readability."""
            if not url.startswith("http"):
                return url
            try:
                from urllib.parse import urlparse
                p     = urlparse(url)
                parts = p.path.strip("/").split("/")
                host  = p.netloc.replace("www.", "")
                slug  = parts[0] if parts else ""
                return f"{host}/{slug}" if slug else host
            except Exception:
                return url[:35]

        # Header
        table_data = [[Paragraph(f"<b>{h}</b>", s_hdr) for h in COLS]]

        # Data rows
        for i, e in enumerate(events, 1):
            status     = _status(e)
            is_ongoing = status == "Ongoing"
            link       = e.get("registration_link", "")
            short_link = _shorten_url(link)

            row = [
                Paragraph(str(i), ParagraphStyle("Idx", parent=s_cell, alignment=TA_CENTER)),
                Paragraph(status, s_ongoing if is_ongoing else s_upcoming),
                Paragraph(e.get("title", "")[:85], s_cell),
                Paragraph(e.get("event_type", ""), s_cell),
                Paragraph(e.get("organizer", "")[:45], s_cell),
                Paragraph(_price_str(e), s_cell),
                Paragraph(e.get("mode", ""), s_cell),
                Paragraph(e.get("location", "")[:45], s_cell),
                Paragraph(e.get("date", ""), s_cell),
                Paragraph(
                    f'<link href="{link}">{short_link}</link>' if link.startswith("http") else (link or "—"),
                    s_link
                ),
            ]
            table_data.append(row)

        # Build alternating + ongoing row styles
        ts_cmds = [
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
            ("TEXTCOLOR",  (0, 0), (-1, 0), HEADER_FG),
            ("LINEBELOW",  (0, 0), (-1, 0), 1.0, BRAND_BLUE),
            # Grid
            ("GRID",          (0, 0), (-1, -1), 0.4, BORDER_COLOR),
            # Padding
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]

        for i, e in enumerate(events, 1):
            row_idx    = i  # 0 = header
            is_ongoing = _status(e) == "Ongoing"
            if is_ongoing:
                ts_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), ROW_ONGOING))
                ts_cmds.append(("LINEAFTER",  (0, row_idx), (0, row_idx), 2.0, GREEN_STATUS))
            else:
                fill = ROW_ODD if i % 2 == 1 else ROW_EVEN
                ts_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), fill))

        tbl = Table(table_data, colWidths=WIDTHS, repeatRows=1, hAlign="LEFT")
        tbl.setStyle(TableStyle(ts_cmds))
        story.append(tbl)

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return filepath
